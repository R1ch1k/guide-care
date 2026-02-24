import csv
import io
import json
import logging
from datetime import date
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import get_async_session
from app.schemas import PatientCreate, PatientOut
from app import crud

logger = logging.getLogger(__name__)

router = APIRouter()


@router.get("", response_model=list[PatientOut])
async def list_patients(db: AsyncSession = Depends(get_async_session)):
    return await crud.list_patients(db)


@router.get("/{patient_id}", response_model=PatientOut)
async def get_patient(patient_id: UUID, db: AsyncSession = Depends(get_async_session)):
    p = await crud.get_patient(db, patient_id)
    if not p:
        raise HTTPException(status_code=404, detail="Patient not found")
    return p


@router.get("/{patient_id}/context")
async def patient_context(patient_id: UUID, db: AsyncSession = Depends(get_async_session)):
    p = await crud.get_patient(db, patient_id)
    if not p:
        raise HTTPException(status_code=404, detail="Patient not found")

    return {
        "id": str(p.id),
        "nhs_number": p.nhs_number,
        "name": f"{p.first_name} {p.last_name}",
        "age": p.age,
        "gender": p.gender,
        "conditions": p.conditions,
        "medications": p.medications,
        "allergies": p.allergies,
        "recent_vitals": p.recent_vitals,
        "clinical_notes": p.clinical_notes,
    }


@router.post("", response_model=PatientOut, status_code=201)
async def create_patient(payload: PatientCreate, db: AsyncSession = Depends(get_async_session)):
    return await crud.create_patient(db, payload)


def _parse_json_field(value: str) -> list | dict:
    """Try to parse a JSON string, return empty list/dict on failure."""
    if not value or not value.strip():
        return []
    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError):
        # Treat comma-separated string as a list
        return [v.strip() for v in value.split(",") if v.strip()]


@router.post("/import", status_code=201)
async def import_patients(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_async_session),
):
    """Import patients from a CSV or Excel file.

    Expected CSV columns:
        nhs_number, first_name, last_name, date_of_birth (YYYY-MM-DD),
        gender, conditions (JSON array or comma-separated),
        medications (JSON array), allergies (JSON array or comma-separated)

    Excel (.xlsx) files use the same column headers in the first row.
    """
    filename = (file.filename or "").lower()
    content = await file.read()

    rows: list[dict] = []

    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append(row)

    elif filename.endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(
                status_code=400,
                detail="Excel support requires openpyxl. Install with: pip install openpyxl",
            )
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({h: (str(v) if v is not None else "") for h, v in zip(headers, row)})

    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Use .csv or .xlsx")

    if not rows:
        raise HTTPException(status_code=400, detail="File contains no data rows")

    imported = 0
    errors: list[str] = []

    for i, row in enumerate(rows, start=1):
        try:
            nhs = row.get("nhs_number", "").strip()
            first = row.get("first_name", "").strip()
            last = row.get("last_name", "").strip()
            dob_str = row.get("date_of_birth", "").strip()
            gender = row.get("gender", "").strip() or None

            if not nhs or not first or not last or not dob_str:
                errors.append(f"Row {i}: missing required field (nhs_number, first_name, last_name, date_of_birth)")
                continue

            dob = date.fromisoformat(dob_str)

            conditions = _parse_json_field(row.get("conditions", ""))
            medications_raw = _parse_json_field(row.get("medications", ""))
            allergies = _parse_json_field(row.get("allergies", ""))

            # Normalize medications to list of {name, dose} dicts
            medications = []
            for m in medications_raw:
                if isinstance(m, dict):
                    medications.append(m)
                elif isinstance(m, str):
                    medications.append({"name": m, "dose": ""})

            payload = PatientCreate(
                nhs_number=nhs,
                first_name=first,
                last_name=last,
                date_of_birth=dob,
                gender=gender,
                conditions=conditions if isinstance(conditions, list) else [],
                medications=[],
                allergies=allergies if isinstance(allergies, list) else [],
            )
            patient = await crud.create_patient(db, payload)

            # Update medications separately (schema uses Medication model)
            if medications:
                patient.medications = medications
                await db.commit()

            imported += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")
            logger.warning("Failed to import row %d: %s", i, e)

    return {
        "imported": imported,
        "total_rows": len(rows),
        "errors": errors[:20],  # Cap error list
    }
